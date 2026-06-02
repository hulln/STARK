#!/usr/bin/env python3
"""Build human-readable side-by-side STARK vs SyntComplex check files per split.

Produces, for the chosen split (dev/train/test):
  manual_side_by_side_<split>.tsv / .csv
  manual_mismatches_only_<split>.tsv
  stark_vs_syntcomplex_manual_check_<split>.xlsx
"""
import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import compare_stark_syntcomplex as compare


BASE_DIR = Path(__file__).resolve().parents[1]


def parse_args():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--split", required=True, choices=["dev", "train", "test"])
    p.add_argument("--corpus", type=Path, default=None)
    p.add_argument("--stark", type=Path, default=None)
    p.add_argument("--stark-details", type=Path, default=None)
    p.add_argument("--syntcomplex", type=Path, default=None)
    p.add_argument("--comparison-summary", type=Path, default=None)
    p.add_argument("--mismatches", type=Path, default=None)
    p.add_argument("--out-dir", type=Path, default=BASE_DIR / "outputs/manual-check")
    args = p.parse_args()
    s = args.split
    args.corpus = args.corpus or BASE_DIR / f"data/sl_ssj-ud-{s}.UD-Slovenian-SSJ-212197d.conllu"
    args.stark = args.stark or BASE_DIR / f"outputs/stark/stark_sl_ssj_{s}.tsv"
    args.stark_details = args.stark_details or BASE_DIR / f"outputs/stark/stark_sl_ssj_{s}_details.tsv"
    args.syntcomplex = args.syntcomplex or BASE_DIR / f"outputs/syntcomplex/syntcomplex_sl_ssj_{s}.tsv"
    args.comparison_summary = args.comparison_summary or BASE_DIR / f"outputs/comparison/comparison_summary_{s}.tsv"
    args.mismatches = args.mismatches or BASE_DIR / f"outputs/comparison/mismatches_{s}.tsv"
    return args


def read_corpus_order(path):
    order = []
    current = {}
    with path.open(encoding="utf-8") as rf:
        for line in rf:
            line = line.rstrip("\n")
            if not line:
                if current.get("sent_id"):
                    order.append(current)
                current = {}
                continue
            if line.startswith("# sent_id = "):
                current["sent_id"] = line.split(" = ", 1)[1]
            elif line.startswith("# text = "):
                current["text"] = line.split(" = ", 1)[1]
        if current.get("sent_id"):
            order.append(current)
    return order


def read_tsv_rows(path):
    with path.open(encoding="utf-8", newline="") as rf:
        return list(csv.DictReader(rf, delimiter="\t"))


def read_summary(path):
    with path.open(encoding="utf-8", newline="") as rf:
        return [row for row in csv.reader(rf, delimiter="\t")]


def fmt_synt(value, kind):
    return compare.format_syntcomplex(value, kind)


def match_display(stark_value, synt_value, kind):
    return stark_value == fmt_synt(synt_value, kind)


def yes_no(value):
    return "yes" if value else "NO"


def build_side_by_side(corpus_order, stark_rows, synt_rows, mismatch_rows):
    mismatch_by_sent_metric = {
        (row["sent_id"], row["metric"]): row for row in mismatch_rows
    }
    headers = [
        "row_no", "sent_id", "sentence",
        "STARK Number of nodes", "SyntComplex # tokens", "nodes match?",
        "STARK MDD", "SyntComplex MDD raw", "SyntComplex MDD rounded like STARK", "MDD match?",
        "STARK NDD", "SyntComplex NDD raw", "SyntComplex NDD rounded like STARK", "NDD match?",
        "STARK Max depth", "SyntComplex MAXIMUM_TREE_DEPTH", "Max depth match?",
        "STARK N clauses", "SyntComplex # clauses", "N clauses match?",
        "STARK N T-units", "SyntComplex # T-units", "N T-units match?",
        "STARK Clauses/T-unit", "SyntComplex Clauses/T-unit raw",
        "SyntComplex Clauses/T-unit rounded like STARK", "Clauses/T-unit match?",
        "STARK example", "STARK tree", "mismatch note",
    ]
    rows = []
    for i, sent in enumerate(corpus_order, 1):
        sent_id = sent["sent_id"]
        stark = stark_rows[sent_id]
        synt = synt_rows[sent_id]
        notes = []
        for metric in ["MDD", "NDD", "Max depth", "N clauses", "N T-units",
                       "Clauses/T-unit", "Number of nodes"]:
            mismatch = mismatch_by_sent_metric.get((sent_id, metric))
            if mismatch:
                notes.append(
                    f"{metric}: STARK={mismatch['stark_value']} vs "
                    f"SyntComplex={mismatch['syntcomplex_value']} "
                    f"({mismatch['mismatch_type']})"
                )
        rows.append([
            i, sent_id, sent.get("text", ""),
            stark["Number of nodes"], synt["#_OF_TOKENS"],
            yes_no(match_display(stark["Number of nodes"], synt["#_OF_TOKENS"], "int")),
            stark["MDD"], synt["MDD"], fmt_synt(synt["MDD"], "float2"),
            yes_no(match_display(stark["MDD"], synt["MDD"], "float2")),
            stark["NDD"], synt["NDD"], fmt_synt(synt["NDD"], "float2"),
            yes_no(match_display(stark["NDD"], synt["NDD"], "float2")),
            stark["Max depth"], synt["MAXIMUM_TREE_DEPTH"],
            yes_no(match_display(stark["Max depth"], synt["MAXIMUM_TREE_DEPTH"], "int")),
            stark["N clauses"], synt["#_OF_CLAUSES"],
            yes_no(match_display(stark["N clauses"], synt["#_OF_CLAUSES"], "float2")),
            stark["N T-units"], synt["#_OF_T-UNITS"],
            yes_no(match_display(stark["N T-units"], synt["#_OF_T-UNITS"], "float2")),
            stark["Clauses/T-unit"], synt["CLAUSES_PER_T-UNIT"],
            fmt_synt(synt["CLAUSES_PER_T-UNIT"], "float2"),
            yes_no(match_display(stark["Clauses/T-unit"], synt["CLAUSES_PER_T-UNIT"], "float2")),
            stark["Example"], stark["Tree"], "; ".join(notes),
        ])
    return headers, rows


def write_delimited(path, headers, rows, delimiter):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as wf:
        writer = csv.writer(wf, delimiter=delimiter)
        writer.writerow(headers)
        writer.writerows(rows)


def add_sheet(wb, title, rows, header=True):
    ws = wb.create_sheet(title)
    for row in rows:
        ws.append(row)
    if header and rows:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    return ws


def autosize(ws, max_width=80):
    for col_idx, column in enumerate(ws.columns, 1):
        width = 8
        for cell in column[:100]:
            if cell.value is None:
                continue
            width = max(width, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def highlight_manual_sheet(ws):
    no_fill = PatternFill("solid", fgColor="F4CCCC")
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    headers = [cell.value for cell in ws[1]]
    note_col = headers.index("mismatch note") + 1
    match_cols = [i + 1 for i, h in enumerate(headers) if h.endswith("match?")]
    for row in ws.iter_rows(min_row=2):
        for col in match_cols:
            if row[col - 1].value == "NO":
                row[col - 1].fill = no_fill
        if row[note_col - 1].value:
            row[note_col - 1].fill = note_fill


def write_workbook(path, manual_headers, manual_rows, summary_rows, mismatch_rows,
                   stark_raw, synt_raw):
    wb = Workbook()
    wb.remove(wb.active)
    manual = add_sheet(wb, "Manual side-by-side", [manual_headers] + manual_rows)
    highlight_manual_sheet(manual)
    add_sheet(wb, "Comparison summary", summary_rows)
    if mismatch_rows:
        add_sheet(wb, "Mismatches",
                  [list(mismatch_rows[0].keys())] + [list(r.values()) for r in mismatch_rows])
    else:
        add_sheet(wb, "Mismatches", [["No mismatches"]], header=False)
    if stark_raw:
        add_sheet(wb, "Raw STARK",
                  [list(stark_raw[0].keys())] + [list(r.values()) for r in stark_raw])
    if synt_raw:
        add_sheet(wb, "Raw SyntComplex",
                  [list(synt_raw[0].keys())] + [list(r.values()) for r in synt_raw])
    for ws in wb.worksheets:
        autosize(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    args = parse_args()
    s = args.split
    key_to_sids, _duplicates = compare.read_stark_details(args.stark_details)
    stark_rows, unmatched = compare.read_stark_rows(args.stark, key_to_sids)
    if unmatched:
        raise RuntimeError(f"Unexpected unmatched STARK tree keys: {len(unmatched)}")
    synt_rows = compare.read_syntcomplex_rows(args.syntcomplex)
    corpus_order = read_corpus_order(args.corpus)
    mismatch_rows = read_tsv_rows(args.mismatches)
    summary_rows = read_summary(args.comparison_summary)
    manual_headers, manual_rows = build_side_by_side(
        corpus_order, stark_rows, synt_rows, mismatch_rows)

    out = args.out_dir
    write_delimited(out / f"manual_side_by_side_{s}.tsv", manual_headers, manual_rows, "\t")
    write_delimited(out / f"manual_side_by_side_{s}.csv", manual_headers, manual_rows, ",")
    write_delimited(
        out / f"manual_mismatches_only_{s}.tsv",
        list(mismatch_rows[0].keys()) if mismatch_rows else ["status"],
        [list(r.values()) for r in mismatch_rows] if mismatch_rows else [["No mismatches"]],
        "\t",
    )
    write_workbook(
        out / f"stark_vs_syntcomplex_manual_check_{s}.xlsx",
        manual_headers, manual_rows, summary_rows, mismatch_rows,
        read_tsv_rows(args.stark), read_tsv_rows(args.syntcomplex),
    )
    for name in (f"manual_side_by_side_{s}.tsv", f"manual_side_by_side_{s}.csv",
                 f"manual_mismatches_only_{s}.tsv",
                 f"stark_vs_syntcomplex_manual_check_{s}.xlsx"):
        print(f"Wrote {out / name}")


if __name__ == "__main__":
    main()
